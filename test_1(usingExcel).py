from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
driver=webdriver.Chrome()
chromoption=Options()
path="franca.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)



for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value)
